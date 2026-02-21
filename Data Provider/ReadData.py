import pandas as pd
import datetime
import configparser
from pandas import read_excel
from openpyxl import load_workbook


def read_config(config_file):
    config_dict = dict()
    parser = configparser.ConfigParser()
    parser.read(config_file)
    sections = parser.sections()
    for section in sections:
        config_dict[section] = dict(parser.items(section))
    return config_dict


def get_exec_list(workbook):
    test_list = []
    df = read_excel(workbook, sheet_name="Execution", engine="openpyxl")
    for ind in df.index:
        if df['Execute'][ind] == "Yes":
            test_list.append(df['Test case'][ind])
    return test_list


def read_master_reg(workbook):
    test_list = get_exec_list(workbook)
    tc_info = dict()
    df = read_excel(workbook, sheet_name="Execution", engine="openpyxl")
    for index, row in df.iterrows():
        if row['Testcase ID'] in test_list:
            tname = row['Test Case Name']
            tfile = row['Test case']
            tc_info[tname] = tfile
    return tc_info


def read_validations(workbook, col):
    validate_dict = dict()
    wb = load_workbook(workbook)
    if 'Validate' in wb.sheetnames:
        df = read_excel(workbook, sheet_name="Validate", engine="openpyxl")
    else:
        return validate_dict
    df = df.replace('\s+', ' ', regex=True)
    for checks in df.itertuples():
        ar_args = []
        if not pd.isna(checks[1]):
            for arg in checks[2:4]:
                if pd.isna(arg):
                    arg = ''
                ar_args.append(arg)
            validate_dict[checks[1]] = ar_args
    return validate_dict


def read_master(workbook):
    test_list = get_exec_list(workbook)
    tc_info = dict()
    df = read_excel(workbook, sheet_name="Execution", engine="openpyxl")
    for index, col in df.iterrows():
        if col['Test case'] in test_list:
            tname = col['Test Case Name']
            tfile = col['Test case']
            tc_info[tname] = tfile
    return tc_info


def read_master_id(workbook):
    tc_info = []
    df = read_excel(workbook, sheet_name="Execution", engine="openpyxl")
    for index, col in df.iterrows():
        if col['Execute'] == 'Yes':
            tc_info.append(col['Test case'])
    return tc_info


def read_exe_master(workbook):
    tc_info = []
    df = read_excel(workbook, sheet_name="Execution", engine="openpyxl")
    for index, col in df.iterrows():
        if col['Execute'] == 'Yes':
            tc_info.append(col['Test case'])
    return tc_info


# Method to read test data from the individual test case files
def read_data(workbook, col):
    ar_data = []
    pages = dict()
    df = read_excel(workbook, sheet_name="Data", engine="openpyxl",)
    df.dropna()
    for index, row in df.iterrows():
        new_dict = row.to_dict()
        if pd.isna(new_dict['Value1']):
            page_name = new_dict['Keys']
            ar_data = []
        else:
            val = new_dict[col]
            if isinstance(val, datetime.date):
                val = val.strftime("%m-%d-%Y")
            ar_data.append(str(val))
            pages[page_name] = ar_data
    return pages


def read_tc_file_rowdata(file, test_list):
    data = dict()
    Xls = pd.ExcelFile(file, engine='openpyxl')
    df1 = pd.read_excel(Xls, 'Data', index_col=0, dtype=str)
    data.update(df1.loc[[test_list]].T.to_dict().items())
    return data


def read_coldata(file, columnname):
    data = dict()
    Xls = pd.ExcelFile(file, engine='openpyxl')
    df1 = pd.read_excel(Xls, 'Data', index_col=0)
    for index, column in df1.items():
        cur_dict = column.to_dict()
        cur_field = dict()
        for key, val in column.to_dict().items():
            if index not in cur_field:
                cur_field[index] = dict()
            if isinstance(val, datetime.date):
                val = val.strftime("%m-%d-%Y")
            cur_field[index].update({str(key): str(val)})
        if index == columnname:
            data = cur_field[index]
    return data


def read_rowdata(file, test_list):
    data = dict()
    Xls = pd.ExcelFile(file, engine='openpyxl')
    df1 = pd.read_excel(Xls, 'Data', index_col=0, dtype=str)
    for tc_id in test_list:
        data.update(df1.loc[[tc_id]].T.to_dict().items())
    return data
